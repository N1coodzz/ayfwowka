import asyncio
import html
import json
import os
import sqlite3
import time
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import aiohttp
from aiohttp import web
from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    CallbackQuery,
    FSInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()

BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
ALLOWED_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "").strip()
TRACKED_WALLET = os.getenv("POLYMARKET_WALLET", "").strip().lower()

POLL_SECONDS = float(os.getenv("POLL_SECONDS", "1.0"))
LIMIT = int(os.getenv("LIMIT", "20"))
MIN_PLAYER_USDC = float(os.getenv("MIN_PLAYER_USDC", "0"))
TAKER_ONLY = os.getenv("TAKER_ONLY", "false").lower() == "true"
FIRST_RUN_SEND_HISTORY = os.getenv("FIRST_RUN_SEND_HISTORY", "false").lower() == "true"

PORT = int(os.getenv("PORT", "8080"))

# Для Railway Volume можно выставить DATA_DIR=/data и примонтировать Volume в /data.
DATA_DIR = Path(os.getenv("DATA_DIR", "."))
DATA_DIR.mkdir(parents=True, exist_ok=True)

DB_PATH = Path(os.getenv("DB_PATH", str(DATA_DIR / "bets.db")))
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", str(DATA_DIR / "bets_report.xlsx")))

POLYMARKET_TRADES_URL = "https://data-api.polymarket.com/trades"

STATUSES = {"OPEN", "WIN", "LOSE", "VOID"}


class RepeatBetFlow(StatesGroup):
    waiting_odds = State()
    waiting_stake = State()
    waiting_edit_odds = State()
    waiting_edit_stake = State()


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def ts_to_iso(ts: Any) -> str:
    try:
        value = float(ts)
        if value > 10_000_000_000:
            value = value / 1000
        return datetime.fromtimestamp(value, tz=timezone.utc).replace(microsecond=0).isoformat()
    except Exception:
        return ""


def as_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        return float(str(value).replace(",", ".").strip())
    except Exception:
        return default


def fmt_money(value: Any) -> str:
    try:
        return f"{float(value):.2f}"
    except Exception:
        return str(value)


def fmt_price(value: Any) -> str:
    try:
        return f"{float(value):.3f}"
    except Exception:
        return str(value)


def h(value: Any) -> str:
    return html.escape(str(value if value is not None else ""))


def validate_config() -> None:
    missing = []
    if not BOT_TOKEN:
        missing.append("TELEGRAM_BOT_TOKEN")
    if not ALLOWED_CHAT_ID:
        missing.append("TELEGRAM_CHAT_ID")
    if not TRACKED_WALLET:
        missing.append("POLYMARKET_WALLET")
    if missing:
        raise RuntimeError(f"Заполни .env: {', '.join(missing)}")
    if not TRACKED_WALLET.startswith("0x") or len(TRACKED_WALLET) != 42:
        raise RuntimeError("POLYMARKET_WALLET должен быть 0x-адресом длиной 42 символа")


def db() -> sqlite3.Connection:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def init_db() -> None:
    with closing(db()) as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS seen_trades (
                trade_uid TEXT PRIMARY KEY,
                first_seen_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trade_uid TEXT UNIQUE NOT NULL,
                detected_at TEXT NOT NULL,
                trade_timestamp TEXT,
                wallet TEXT NOT NULL,
                side TEXT,
                outcome TEXT,
                price REAL,
                size REAL,
                usdc_size REAL,
                market_title TEXT,
                market_url TEXT,
                tx_hash TEXT,
                raw_json TEXT
            );

            CREATE TABLE IF NOT EXISTS bets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                alert_id INTEGER,
                created_at TEXT NOT NULL,
                market_title TEXT,
                market_url TEXT,
                player_side TEXT,
                player_outcome TEXT,
                player_price REAL,
                player_usdc_size REAL,
                my_odds REAL NOT NULL,
                my_stake REAL NOT NULL,
                status TEXT NOT NULL DEFAULT 'OPEN',
                profit REAL NOT NULL DEFAULT 0,
                tx_hash TEXT,
                notes TEXT,
                FOREIGN KEY(alert_id) REFERENCES alerts(id)
            );
            """
        )
        con.commit()


def allowed_user(message: Message) -> bool:
    return str(message.chat.id) == ALLOWED_CHAT_ID or (
        message.from_user is not None and str(message.from_user.id) == ALLOWED_CHAT_ID
    )


def allowed_callback(callback: CallbackQuery) -> bool:
    return callback.from_user is not None and str(callback.from_user.id) == ALLOWED_CHAT_ID


def trade_uid(t: Dict[str, Any]) -> str:
    return "|".join(
        [
            str(t.get("transactionHash", "")),
            str(t.get("conditionId", "")),
            str(t.get("asset", "")),
            str(t.get("side", "")),
            str(t.get("outcomeIndex", "")),
            str(t.get("size", "")),
            str(t.get("price", "")),
            str(t.get("timestamp", "")),
        ]
    )


def market_url(t: Dict[str, Any]) -> str:
    event_slug = t.get("eventSlug")
    slug = t.get("slug")
    if event_slug:
        return f"https://polymarket.com/event/{event_slug}"
    if slug:
        return f"https://polymarket.com/event/{slug}"
    return "https://polymarket.com"


def calc_usdc(t: Dict[str, Any]) -> float:
    if t.get("usdcSize") is not None:
        return as_float(t.get("usdcSize"))
    return as_float(t.get("size")) * as_float(t.get("price"))


def insert_seen(uid: str) -> None:
    with closing(db()) as con:
        con.execute(
            "INSERT OR IGNORE INTO seen_trades(trade_uid, first_seen_at) VALUES(?, ?)",
            (uid, now_iso()),
        )
        con.commit()


def is_seen(uid: str) -> bool:
    with closing(db()) as con:
        row = con.execute("SELECT 1 FROM seen_trades WHERE trade_uid = ?", (uid,)).fetchone()
        return row is not None


def seen_count() -> int:
    with closing(db()) as con:
        row = con.execute("SELECT COUNT(*) AS c FROM seen_trades").fetchone()
        return int(row["c"])


def insert_alert_from_trade(t: Dict[str, Any]) -> int:
    uid = trade_uid(t)
    with closing(db()) as con:
        con.execute(
            """
            INSERT OR IGNORE INTO alerts(
                trade_uid, detected_at, trade_timestamp, wallet, side, outcome,
                price, size, usdc_size, market_title, market_url, tx_hash, raw_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                uid,
                now_iso(),
                ts_to_iso(t.get("timestamp")),
                TRACKED_WALLET,
                str(t.get("side", "")).upper(),
                str(t.get("outcome", "")),
                as_float(t.get("price")),
                as_float(t.get("size")),
                calc_usdc(t),
                str(t.get("title", "")),
                market_url(t),
                str(t.get("transactionHash", "")),
                json.dumps(t, ensure_ascii=False),
            ),
        )
        row = con.execute("SELECT id FROM alerts WHERE trade_uid = ?", (uid,)).fetchone()
        con.commit()
        return int(row["id"])


def create_test_alert() -> int:
    uid = f"test-{int(time.time())}"
    with closing(db()) as con:
        con.execute(
            """
            INSERT INTO alerts(
                trade_uid, detected_at, trade_timestamp, wallet, side, outcome,
                price, size, usdc_size, market_title, market_url, tx_hash, raw_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                uid,
                now_iso(),
                now_iso(),
                TRACKED_WALLET,
                "BUY",
                "YES",
                0.45,
                100,
                45,
                "TEST MARKET — проверка кнопки Повторить",
                "https://polymarket.com",
                "test",
                "{}",
            ),
        )
        alert_id = int(con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
        con.commit()
        return alert_id


def get_alert(alert_id: int) -> Optional[sqlite3.Row]:
    with closing(db()) as con:
        return con.execute("SELECT * FROM alerts WHERE id = ?", (alert_id,)).fetchone()


def create_bet(alert_id: int, odds: float, stake: float) -> int:
    alert = get_alert(alert_id)
    if not alert:
        raise RuntimeError("Алерт не найден")

    with closing(db()) as con:
        con.execute(
            """
            INSERT INTO bets(
                alert_id, created_at, market_title, market_url, player_side,
                player_outcome, player_price, player_usdc_size, my_odds,
                my_stake, status, profit, tx_hash
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'OPEN', 0, ?)
            """,
            (
                alert_id,
                now_iso(),
                alert["market_title"],
                alert["market_url"],
                alert["side"],
                alert["outcome"],
                alert["price"],
                alert["usdc_size"],
                odds,
                stake,
                alert["tx_hash"],
            ),
        )
        bet_id = int(con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"])
        con.commit()
        return bet_id


def calc_profit(status: str, odds: float, stake: float) -> float:
    if status == "WIN":
        return round(stake * (odds - 1), 2)
    if status == "LOSE":
        return round(-stake, 2)
    return 0.0


def update_bet_status(bet_id: int, status: str) -> None:
    if status not in STATUSES:
        raise RuntimeError("Некорректный статус")
    with closing(db()) as con:
        bet = con.execute("SELECT * FROM bets WHERE id = ?", (bet_id,)).fetchone()
        if not bet:
            raise RuntimeError("Ставка не найдена")
        profit = calc_profit(status, float(bet["my_odds"]), float(bet["my_stake"]))
        con.execute("UPDATE bets SET status = ?, profit = ? WHERE id = ?", (status, profit, bet_id))
        con.commit()


def update_bet_field(bet_id: int, field: str, value: float) -> None:
    if field not in {"my_odds", "my_stake"}:
        raise RuntimeError("Некорректное поле")
    with closing(db()) as con:
        bet = con.execute("SELECT * FROM bets WHERE id = ?", (bet_id,)).fetchone()
        if not bet:
            raise RuntimeError("Ставка не найдена")
        new_odds = value if field == "my_odds" else float(bet["my_odds"])
        new_stake = value if field == "my_stake" else float(bet["my_stake"])
        profit = calc_profit(str(bet["status"]), new_odds, new_stake)
        con.execute(
            f"UPDATE bets SET {field} = ?, profit = ? WHERE id = ?",
            (value, profit, bet_id),
        )
        con.commit()


def get_bet(bet_id: int) -> Optional[sqlite3.Row]:
    with closing(db()) as con:
        return con.execute("SELECT * FROM bets WHERE id = ?", (bet_id,)).fetchone()


def get_bets(status: Optional[str] = None, limit: int = 10, offset: int = 0) -> List[sqlite3.Row]:
    with closing(db()) as con:
        if status and status in STATUSES:
            rows = con.execute(
                "SELECT * FROM bets WHERE status = ? ORDER BY id DESC LIMIT ? OFFSET ?",
                (status, limit, offset),
            ).fetchall()
        else:
            rows = con.execute(
                "SELECT * FROM bets ORDER BY id DESC LIMIT ? OFFSET ?",
                (limit, offset),
            ).fetchall()
        return rows


def get_stats(status_filter: Optional[str] = None) -> Dict[str, Any]:
    with closing(db()) as con:
        where = ""
        params: Tuple[Any, ...] = ()
        if status_filter and status_filter in STATUSES:
            where = "WHERE status = ?"
            params = (status_filter,)

        rows = con.execute(f"SELECT * FROM bets {where}", params).fetchall()

    total = len(rows)
    open_bets = sum(1 for r in rows if r["status"] == "OPEN")
    closed = [r for r in rows if r["status"] in {"WIN", "LOSE"}]
    wins = sum(1 for r in rows if r["status"] == "WIN")
    loses = sum(1 for r in rows if r["status"] == "LOSE")
    voids = sum(1 for r in rows if r["status"] == "VOID")
    stake_total = sum(float(r["my_stake"]) for r in rows)
    settled_stake = sum(float(r["my_stake"]) for r in rows if r["status"] in {"WIN", "LOSE"})
    profit = sum(float(r["profit"]) for r in rows)
    avg_odds = sum(float(r["my_odds"]) for r in rows) / total if total else 0
    winrate = wins / len(closed) * 100 if closed else 0
    roi = profit / settled_stake * 100 if settled_stake else 0

    return {
        "total": total,
        "open": open_bets,
        "wins": wins,
        "loses": loses,
        "voids": voids,
        "stake_total": stake_total,
        "settled_stake": settled_stake,
        "profit": profit,
        "avg_odds": avg_odds,
        "winrate": winrate,
        "roi": roi,
    }


def group_stats(field: str) -> List[Dict[str, Any]]:
    if field not in {"market_title", "player_outcome"}:
        raise RuntimeError("Некорректная группировка")

    with closing(db()) as con:
        rows = con.execute(f"SELECT * FROM bets ORDER BY id DESC").fetchall()

    groups: Dict[str, List[sqlite3.Row]] = {}
    for r in rows:
        key = str(r[field] or "Unknown")
        groups.setdefault(key, []).append(r)

    result = []
    for key, items in groups.items():
        closed = [r for r in items if r["status"] in {"WIN", "LOSE"}]
        wins = sum(1 for r in items if r["status"] == "WIN")
        loses = sum(1 for r in items if r["status"] == "LOSE")
        stake = sum(float(r["my_stake"]) for r in items)
        settled_stake = sum(float(r["my_stake"]) for r in items if r["status"] in {"WIN", "LOSE"})
        profit = sum(float(r["profit"]) for r in items)
        result.append(
            {
                "name": key,
                "bets": len(items),
                "open": sum(1 for r in items if r["status"] == "OPEN"),
                "wins": wins,
                "loses": loses,
                "voids": sum(1 for r in items if r["status"] == "VOID"),
                "stake": stake,
                "profit": profit,
                "roi": profit / settled_stake * 100 if settled_stake else 0,
                "winrate": wins / len(closed) * 100 if closed else 0,
                "avg_odds": sum(float(r["my_odds"]) for r in items) / len(items) if items else 0,
            }
        )

    result.sort(key=lambda x: (x["profit"], x["bets"]), reverse=True)
    return result


def status_emoji(status: str) -> str:
    return {"OPEN": "🟡", "WIN": "✅", "LOSE": "❌", "VOID": "↩️"}.get(status, "⚪️")


def repeat_keyboard(alert_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔁 Повторить", callback_data=f"repeat:{alert_id}")]
        ]
    )


def bet_keyboard(bet_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Win", callback_data=f"status:{bet_id}:WIN"),
                InlineKeyboardButton(text="❌ Lose", callback_data=f"status:{bet_id}:LOSE"),
                InlineKeyboardButton(text="↩️ Void", callback_data=f"status:{bet_id}:VOID"),
            ],
            [
                InlineKeyboardButton(text="📝 Кф", callback_data=f"editodds:{bet_id}"),
                InlineKeyboardButton(text="💵 Сумма", callback_data=f"editstake:{bet_id}"),
            ],
        ]
    )


def filters_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Все", callback_data="list:ALL:0"),
                InlineKeyboardButton(text="Открытые", callback_data="list:OPEN:0"),
            ],
            [
                InlineKeyboardButton(text="Win", callback_data="list:WIN:0"),
                InlineKeyboardButton(text="Lose", callback_data="list:LOSE:0"),
                InlineKeyboardButton(text="Void", callback_data="list:VOID:0"),
            ],
            [InlineKeyboardButton(text="📊 Экспорт Excel", callback_data="export")],
        ]
    )


def build_alert_text(alert: sqlite3.Row, detected_delay: Optional[float] = None) -> str:
    side = str(alert["side"] or "?").upper()
    emoji = "🟢" if side == "BUY" else "🔴" if side == "SELL" else "⚪️"
    delay = f"\n⚡ Задержка по timestamp: ~{detected_delay:.1f}s" if detected_delay is not None else ""
    return (
        f"{emoji} <b>{h(side)} {h(alert['outcome'])}</b> @ <b>{fmt_price(alert['price'])}</b>\n"
        f"💵 Игрок: ~<b>{fmt_money(alert['usdc_size'])} USDC</b> | 🎟 {fmt_money(alert['size'])}\n"
        f"📌 {h(alert['market_title'])}\n"
        f"🔗 {h(alert['market_url'])}\n"
        f"TX: <code>{h(str(alert['tx_hash'])[:10] + '...' + str(alert['tx_hash'])[-6:] if alert['tx_hash'] else '—')}</code>"
        f"{delay}"
    )


def build_bet_text(bet: sqlite3.Row) -> str:
    return (
        f"{status_emoji(bet['status'])} <b>Ставка #{bet['id']}</b> — <b>{h(bet['status'])}</b>\n"
        f"📌 {h(bet['market_title'])}\n"
        f"🎯 Повтор: <b>{h(bet['player_side'])} {h(bet['player_outcome'])}</b>\n"
        f"📈 Кф: <b>{fmt_price(bet['my_odds'])}</b>\n"
        f"💵 Сумма: <b>{fmt_money(bet['my_stake'])}</b>\n"
        f"💰 P/L: <b>{fmt_money(bet['profit'])}</b>\n"
        f"🔗 {h(bet['market_url'])}"
    )


def build_stats_text() -> str:
    s = get_stats()
    return (
        f"📊 <b>Статистика</b>\n\n"
        f"Всего ставок: <b>{s['total']}</b>\n"
        f"Открытых: <b>{s['open']}</b>\n"
        f"Win / Lose / Void: <b>{s['wins']} / {s['loses']} / {s['voids']}</b>\n"
        f"Winrate: <b>{s['winrate']:.1f}%</b>\n"
        f"Оборот: <b>{s['stake_total']:.2f}</b>\n"
        f"Закрытый оборот: <b>{s['settled_stake']:.2f}</b>\n"
        f"Профит: <b>{s['profit']:+.2f}</b>\n"
        f"ROI: <b>{s['roi']:+.1f}%</b>\n"
        f"Средний кф: <b>{s['avg_odds']:.2f}</b>"
    )


def parse_positive_float(text: str) -> Optional[float]:
    text = text.strip().replace(",", ".")
    try:
        value = float(text)
        if value <= 0:
            return None
        return value
    except Exception:
        return None


def trade_delay_seconds(t: Dict[str, Any]) -> Optional[float]:
    try:
        ts = float(t.get("timestamp"))
        if ts > 10_000_000_000:
            ts = ts / 1000
        return max(0.0, time.time() - ts)
    except Exception:
        return None


async def fetch_trades(session: aiohttp.ClientSession) -> List[Dict[str, Any]]:
    params = {
        "user": TRACKED_WALLET,
        "limit": LIMIT,
        "offset": 0,
        "takerOnly": str(TAKER_ONLY).lower(),
    }
    async with session.get(
        POLYMARKET_TRADES_URL,
        params=params,
        timeout=aiohttp.ClientTimeout(total=6),
    ) as response:
        body = await response.text()
        if response.status != 200:
            raise RuntimeError(f"Polymarket {response.status}: {body[:300]}")
        data = json.loads(body)
        if not isinstance(data, list):
            raise RuntimeError(f"Unexpected Polymarket response: {data}")
        return data


async def trade_watcher(bot: Bot) -> None:
    first_loop = seen_count() == 0
    connector = aiohttp.TCPConnector(limit=10, ttl_dns_cache=300, keepalive_timeout=30)
    headers = {
        "User-Agent": "polymarket-tracker-report-bot/1.0",
        "Accept": "application/json",
    }

    async with aiohttp.ClientSession(connector=connector, headers=headers) as session:
        while True:
            started = time.perf_counter()
            try:
                trades = await fetch_trades(session)
                fresh: List[Tuple[int, Dict[str, Any]]] = []

                for t in trades:
                    if calc_usdc(t) < MIN_PLAYER_USDC:
                        continue

                    uid = trade_uid(t)
                    if not is_seen(uid):
                        alert_id = insert_alert_from_trade(t)
                        insert_seen(uid)
                        fresh.append((alert_id, t))

                if first_loop and not FIRST_RUN_SEND_HISTORY:
                    first_loop = False
                else:
                    fresh.sort(key=lambda item: float(item[1].get("timestamp") or 0))
                    for alert_id, t in fresh:
                        alert = get_alert(alert_id)
                        if alert:
                            await bot.send_message(
                                ALLOWED_CHAT_ID,
                                build_alert_text(alert, trade_delay_seconds(t)),
                                reply_markup=repeat_keyboard(alert_id),
                                disable_web_page_preview=True,
                            )
                    first_loop = False

            except Exception as e:
                print(f"[watcher error] {e}")

            elapsed = time.perf_counter() - started
            await asyncio.sleep(max(0.05, POLL_SECONDS - elapsed))


async def cmd_start(message: Message) -> None:
    if not allowed_user(message):
        return
    await message.answer(
        "✅ Бот запущен.\n\n"
        f"👛 Слежу за кошельком:\n<code>{h(TRACKED_WALLET)}</code>\n"
        f"⏱ Проверка: <b>{POLL_SECONDS}s</b>\n"
        f"🎯 takerOnly: <b>{str(TAKER_ONLY).lower()}</b>\n\n"
        "Команды:\n"
        "/ping_polymarket — проверить реальные сделки кошелька\n"
        "/test_alert — проверить сценарий без реальной сделки\n"
        "/stats — статистика\n"
        "/bets — ставки и фильтры\n"
        "/open — открытые ставки\n"
        "/markets — статистика по рынкам\n"
        "/export — Excel-отчёт"
    )


async def cmd_test_alert(message: Message) -> None:
    if not allowed_user(message):
        return
    alert_id = create_test_alert()
    alert = get_alert(alert_id)
    await message.answer(
        build_alert_text(alert),
        reply_markup=repeat_keyboard(alert_id),
        disable_web_page_preview=True,
    )


async def cmd_ping_polymarket(message: Message, bot: Bot) -> None:
    """Отправляет последние реальные сделки кошелька — проверка связки Polymarket -> Telegram."""
    if not allowed_user(message):
        return

    await message.answer("🔎 Проверяю последние реальные сделки кошелька Polymarket...")

    try:
        async with aiohttp.ClientSession(headers={"Accept": "application/json"}) as session:
            trades = await fetch_trades(session)

        trades = [t for t in trades if calc_usdc(t) >= MIN_PLAYER_USDC]

        if not trades:
            await message.answer("Реальных сделок по этому кошельку не найдено в ответе API.")
            return

        # До 3 последних сделок. Это реальные алерты с кнопкой «Повторить».
        for t in reversed(trades[:3]):
            alert_id = insert_alert_from_trade(t)
            alert = get_alert(alert_id)
            if alert:
                await message.answer(
                    build_alert_text(alert, trade_delay_seconds(t)),
                    reply_markup=repeat_keyboard(alert_id),
                    disable_web_page_preview=True,
                )

        await message.answer(
            "✅ Связка работает: Polymarket → этот кошелёк → твой Telegram.\n"
            "Выше отправил последние реальные сделки с кнопкой «Повторить»."
        )

    except Exception as e:
        await message.answer(f"❌ Ошибка проверки Polymarket:\n<code>{h(e)}</code>")


async def cmd_stats(message: Message) -> None:
    if not allowed_user(message):
        return
    await message.answer(build_stats_text(), reply_markup=filters_keyboard())


async def send_bets_list(message_or_callback: Any, status: Optional[str] = None, offset: int = 0) -> None:
    rows = get_bets(status if status != "ALL" else None, limit=10, offset=offset)
    title = "Все ставки" if not status or status == "ALL" else f"Ставки: {status}"

    if not rows:
        text = f"Пока нет ставок в фильтре: <b>{h(title)}</b>"
    else:
        lines = [f"📋 <b>{h(title)}</b>"]
        for r in rows:
            lines.append(
                f"\n{status_emoji(r['status'])} <b>#{r['id']}</b> "
                f"{h(r['player_side'])} {h(r['player_outcome'])} | "
                f"кф {fmt_price(r['my_odds'])} | "
                f"{fmt_money(r['my_stake'])} | "
                f"P/L {float(r['profit']):+.2f}\n"
                f"{h(str(r['market_title'])[:80])}"
            )
        text = "\n".join(lines)

    keyboard_rows = []
    for r in rows[:8]:
        keyboard_rows.append([InlineKeyboardButton(text=f"{status_emoji(r['status'])} Ставка #{r['id']}", callback_data=f"bet:{r['id']}")])

    nav = []
    if offset >= 10:
        nav.append(InlineKeyboardButton(text="⬅️ Назад", callback_data=f"list:{status or 'ALL'}:{max(0, offset-10)}"))
    if len(rows) == 10:
        nav.append(InlineKeyboardButton(text="➡️ Далее", callback_data=f"list:{status or 'ALL'}:{offset+10}"))
    if nav:
        keyboard_rows.append(nav)

    keyboard_rows.append(
        [
            InlineKeyboardButton(text="Все", callback_data="list:ALL:0"),
            InlineKeyboardButton(text="Open", callback_data="list:OPEN:0"),
            InlineKeyboardButton(text="Win", callback_data="list:WIN:0"),
            InlineKeyboardButton(text="Lose", callback_data="list:LOSE:0"),
        ]
    )
    keyboard_rows.append([InlineKeyboardButton(text="📊 Экспорт Excel", callback_data="export")])
    keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_rows)

    if isinstance(message_or_callback, CallbackQuery):
        await message_or_callback.message.edit_text(text, reply_markup=keyboard, disable_web_page_preview=True)
    else:
        await message_or_callback.answer(text, reply_markup=keyboard, disable_web_page_preview=True)


async def cmd_bets(message: Message) -> None:
    if not allowed_user(message):
        return
    await send_bets_list(message, "ALL", 0)


async def cmd_open(message: Message) -> None:
    if not allowed_user(message):
        return
    await send_bets_list(message, "OPEN", 0)


async def cmd_markets(message: Message) -> None:
    if not allowed_user(message):
        return
    groups = group_stats("market_title")[:10]
    if not groups:
        await message.answer("Пока нет ставок для статистики по рынкам.")
        return

    lines = ["🏟 <b>Статистика по рынкам</b>"]
    for g in groups:
        lines.append(
            f"\n<b>{h(str(g['name'])[:80])}</b>\n"
            f"Ставок: {g['bets']} | W/L/V/O: {g['wins']}/{g['loses']}/{g['voids']}/{g['open']}\n"
            f"Winrate: {g['winrate']:.1f}% | ROI: {g['roi']:+.1f}% | P/L: {g['profit']:+.2f}"
        )
    await message.answer("\n".join(lines))


async def cmd_export(message: Message) -> None:
    if not allowed_user(message):
        return
    path = export_excel()
    await message.answer_document(FSInputFile(path), caption="📊 Excel-отчёт обновлён")


async def cb_repeat(callback: CallbackQuery, state: FSMContext) -> None:
    if not allowed_callback(callback):
        return
    alert_id = int(callback.data.split(":")[1])
    alert = get_alert(alert_id)
    if not alert:
        await callback.answer("Алерт не найден", show_alert=True)
        return
    await state.set_state(RepeatBetFlow.waiting_odds)
    await state.update_data(alert_id=alert_id)
    await callback.message.answer(
        "🔁 Записываем повтор.\n\n"
        f"📌 {h(alert['market_title'])}\n"
        f"🎯 Игрок: <b>{h(alert['side'])} {h(alert['outcome'])}</b> @ {fmt_price(alert['price'])}\n\n"
        "Какой коэффициент ты взял? Например: <code>2.15</code>"
    )
    await callback.answer()


async def process_odds(message: Message, state: FSMContext) -> None:
    if not allowed_user(message):
        return
    odds = parse_positive_float(message.text or "")
    if odds is None or odds < 1.01:
        await message.answer("Кф должен быть числом больше 1.01. Пример: <code>2.15</code>")
        return
    await state.update_data(odds=odds)
    await state.set_state(RepeatBetFlow.waiting_stake)
    await message.answer("Какую сумму поставил? Например: <code>25</code>")


async def process_stake(message: Message, state: FSMContext) -> None:
    if not allowed_user(message):
        return
    stake = parse_positive_float(message.text or "")
    if stake is None:
        await message.answer("Сумма должна быть положительным числом. Пример: <code>25</code>")
        return

    data = await state.get_data()
    alert_id = int(data["alert_id"])
    odds = float(data["odds"])
    bet_id = create_bet(alert_id, odds, stake)
    await state.clear()

    bet = get_bet(bet_id)
    await message.answer(
        "✅ Ставка записана.\n\n" + build_bet_text(bet),
        reply_markup=bet_keyboard(bet_id),
        disable_web_page_preview=True,
    )


async def cb_status(callback: CallbackQuery) -> None:
    if not allowed_callback(callback):
        return
    _, bet_id_str, status = callback.data.split(":")
    bet_id = int(bet_id_str)
    update_bet_status(bet_id, status)
    bet = get_bet(bet_id)
    await callback.message.edit_text(
        build_bet_text(bet),
        reply_markup=bet_keyboard(bet_id),
        disable_web_page_preview=True,
    )
    await callback.answer(f"Статус: {status}")


async def cb_bet(callback: CallbackQuery) -> None:
    if not allowed_callback(callback):
        return
    bet_id = int(callback.data.split(":")[1])
    bet = get_bet(bet_id)
    if not bet:
        await callback.answer("Ставка не найдена", show_alert=True)
        return
    await callback.message.answer(
        build_bet_text(bet),
        reply_markup=bet_keyboard(bet_id),
        disable_web_page_preview=True,
    )
    await callback.answer()


async def cb_list(callback: CallbackQuery) -> None:
    if not allowed_callback(callback):
        return
    _, status, offset = callback.data.split(":")
    await send_bets_list(callback, status, int(offset))
    await callback.answer()


async def cb_export(callback: CallbackQuery) -> None:
    if not allowed_callback(callback):
        return
    path = export_excel()
    await callback.message.answer_document(FSInputFile(path), caption="📊 Excel-отчёт обновлён")
    await callback.answer()


async def cb_edit_odds(callback: CallbackQuery, state: FSMContext) -> None:
    if not allowed_callback(callback):
        return
    bet_id = int(callback.data.split(":")[1])
    await state.set_state(RepeatBetFlow.waiting_edit_odds)
    await state.update_data(bet_id=bet_id)
    await callback.message.answer(f"Введи новый кф для ставки #{bet_id}.")
    await callback.answer()


async def cb_edit_stake(callback: CallbackQuery, state: FSMContext) -> None:
    if not allowed_callback(callback):
        return
    bet_id = int(callback.data.split(":")[1])
    await state.set_state(RepeatBetFlow.waiting_edit_stake)
    await state.update_data(bet_id=bet_id)
    await callback.message.answer(f"Введи новую сумму для ставки #{bet_id}.")
    await callback.answer()


async def process_edit_odds(message: Message, state: FSMContext) -> None:
    if not allowed_user(message):
        return
    odds = parse_positive_float(message.text or "")
    if odds is None or odds < 1.01:
        await message.answer("Кф должен быть числом больше 1.01.")
        return
    data = await state.get_data()
    bet_id = int(data["bet_id"])
    update_bet_field(bet_id, "my_odds", odds)
    await state.clear()
    bet = get_bet(bet_id)
    await message.answer("✅ Кф обновлён.\n\n" + build_bet_text(bet), reply_markup=bet_keyboard(bet_id))


async def process_edit_stake(message: Message, state: FSMContext) -> None:
    if not allowed_user(message):
        return
    stake = parse_positive_float(message.text or "")
    if stake is None:
        await message.answer("Сумма должна быть положительным числом.")
        return
    data = await state.get_data()
    bet_id = int(data["bet_id"])
    update_bet_field(bet_id, "my_stake", stake)
    await state.clear()
    bet = get_bet(bet_id)
    await message.answer("✅ Сумма обновлена.\n\n" + build_bet_text(bet), reply_markup=bet_keyboard(bet_id))


def export_excel() -> Path:
    with closing(db()) as con:
        bets = con.execute("SELECT * FROM bets ORDER BY id ASC").fetchall()
        alerts = con.execute("SELECT * FROM alerts ORDER BY id ASC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bets"

    headers = [
        "ID", "Created UTC", "Status", "Stake", "Odds", "Profit", "ROI %",
        "Player Side", "Outcome", "Player Price", "Player USDC",
        "Market", "URL", "TX"
    ]
    ws.append(headers)

    for r in bets:
        stake = float(r["my_stake"])
        roi = float(r["profit"]) / stake * 100 if stake and r["status"] in {"WIN", "LOSE"} else 0
        ws.append([
            r["id"], r["created_at"], r["status"], stake, float(r["my_odds"]),
            float(r["profit"]), roi, r["player_side"], r["player_outcome"],
            float(r["player_price"] or 0), float(r["player_usdc_size"] or 0),
            r["market_title"], r["market_url"], r["tx_hash"]
        ])

    stats_ws = wb.create_sheet("Stats")
    s = get_stats()
    stats_rows = [
        ("Total bets", s["total"]),
        ("Open bets", s["open"]),
        ("Wins", s["wins"]),
        ("Loses", s["loses"]),
        ("Voids", s["voids"]),
        ("Turnover", round(s["stake_total"], 2)),
        ("Settled turnover", round(s["settled_stake"], 2)),
        ("Profit", round(s["profit"], 2)),
        ("ROI %", round(s["roi"], 2)),
        ("Winrate %", round(s["winrate"], 2)),
        ("Average odds", round(s["avg_odds"], 3)),
    ]
    stats_ws.append(["Metric", "Value"])
    for row in stats_rows:
        stats_ws.append(list(row))

    markets_ws = wb.create_sheet("By Market")
    markets_ws.append(["Market", "Bets", "Open", "Wins", "Loses", "Voids", "Stake", "Profit", "ROI %", "Winrate %", "Avg Odds"])
    for g in group_stats("market_title"):
        markets_ws.append([
            g["name"], g["bets"], g["open"], g["wins"], g["loses"], g["voids"],
            round(g["stake"], 2), round(g["profit"], 2), round(g["roi"], 2),
            round(g["winrate"], 2), round(g["avg_odds"], 3)
        ])

    outcomes_ws = wb.create_sheet("By Outcome")
    outcomes_ws.append(["Outcome", "Bets", "Open", "Wins", "Loses", "Voids", "Stake", "Profit", "ROI %", "Winrate %", "Avg Odds"])
    for g in group_stats("player_outcome"):
        outcomes_ws.append([
            g["name"], g["bets"], g["open"], g["wins"], g["loses"], g["voids"],
            round(g["stake"], 2), round(g["profit"], 2), round(g["roi"], 2),
            round(g["winrate"], 2), round(g["avg_odds"], 3)
        ])

    alerts_ws = wb.create_sheet("Alerts")
    alerts_ws.append([
        "ID", "Detected UTC", "Trade UTC", "Side", "Outcome", "Price",
        "Size", "USDC", "Market", "URL", "TX"
    ])
    for r in alerts:
        alerts_ws.append([
            r["id"], r["detected_at"], r["trade_timestamp"], r["side"], r["outcome"],
            r["price"], r["size"], r["usdc_size"], r["market_title"], r["market_url"], r["tx_hash"]
        ])

    style_workbook(wb)
    wb.save(EXCEL_PATH)
    return EXCEL_PATH


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 42)

        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 10)

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20

        for col in ["D", "E", "F", "G", "J", "K"]:
            if col in ws.column_dimensions:
                for cell in ws[col][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"


def register_handlers(dp: Dispatcher) -> None:
    dp.message.register(cmd_start, Command("start"))
    dp.message.register(cmd_start, Command("help"))
    dp.message.register(cmd_ping_polymarket, Command("ping_polymarket"))
    dp.message.register(cmd_test_alert, Command("test_alert"))
    dp.message.register(cmd_stats, Command("stats"))
    dp.message.register(cmd_bets, Command("bets"))
    dp.message.register(cmd_open, Command("open"))
    dp.message.register(cmd_markets, Command("markets"))
    dp.message.register(cmd_export, Command("export"))

    dp.callback_query.register(cb_repeat, F.data.startswith("repeat:"))
    dp.callback_query.register(cb_status, F.data.startswith("status:"))
    dp.callback_query.register(cb_bet, F.data.startswith("bet:"))
    dp.callback_query.register(cb_list, F.data.startswith("list:"))
    dp.callback_query.register(cb_export, F.data == "export")
    dp.callback_query.register(cb_edit_odds, F.data.startswith("editodds:"))
    dp.callback_query.register(cb_edit_stake, F.data.startswith("editstake:"))

    dp.message.register(process_odds, RepeatBetFlow.waiting_odds)
    dp.message.register(process_stake, RepeatBetFlow.waiting_stake)
    dp.message.register(process_edit_odds, RepeatBetFlow.waiting_edit_odds)
    dp.message.register(process_edit_stake, RepeatBetFlow.waiting_edit_stake)


async def health_root(request: web.Request) -> web.Response:
    return web.json_response({
        "ok": True,
        "service": "polymarket-telegram-tracker",
        "wallet": TRACKED_WALLET,
        "poll_seconds": POLL_SECONDS,
        "db_path": str(DB_PATH),
    })


async def health_stats(request: web.Request) -> web.Response:
    return web.json_response(get_stats())


async def start_web_server() -> None:
    app = web.Application()
    app.router.add_get("/", health_root)
    app.router.add_get("/health", health_root)
    app.router.add_get("/stats.json", health_stats)

    runner = web.AppRunner(app)
    await runner.setup()

    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()

    print(f"[web] started on 0.0.0.0:{PORT}")


async def main() -> None:
    validate_config()
    init_db()

    bot = Bot(
        token=BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )

    # На Railway используем long polling. Если раньше был webhook — удаляем.
    await bot.delete_webhook(drop_pending_updates=True)

    dp = Dispatcher(storage=MemoryStorage())
    register_handlers(dp)

    await start_web_server()

    await bot.send_message(
        ALLOWED_CHAT_ID,
        "✅ Polymarket tracker запущен.\n"
        f"👛 Wallet: <code>{h(TRACKED_WALLET)}</code>\n"
        f"⏱ Poll: <b>{POLL_SECONDS}s</b>\n"
        "Проверка реального кошелька: /ping_polymarket",
    )

    asyncio.create_task(trade_watcher(bot))
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
